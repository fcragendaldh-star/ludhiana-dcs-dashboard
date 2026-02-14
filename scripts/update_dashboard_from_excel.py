#!/usr/bin/env python3
"""
Update index.html from a DCS status Excel sheet.

Usage:
    python scripts/update_dashboard_from_excel.py <excel_path>
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CANONICAL_ORDER = [
    "Raikot",
    "Khanna",
    "Ludhiana (East)",
    "Payal",
    "Jagraon",
    "Samrala",
    "Ludhiana (West)",
]


DISPLAY_NAME_MAP = {
    "raikot": "Raikot",
    "khanna": "Khanna",
    "payal": "Payal",
    "jagraon": "Jagraon",
    "samrala": "Samrala",
    "ludhianaeast": "Ludhiana (East)",
    "ludhianawest": "Ludhiana (West)",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip().lower()


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def display_name(raw_name: str) -> str:
    return DISPLAY_NAME_MAP.get(normalize_name(raw_name), raw_name.strip())


def parse_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text == "-":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def to_int(value: Any) -> int:
    return int(round(parse_number(value)))


def to_percent(value: Any) -> float:
    return round(parse_number(value), 2)


def format_indian_number(value: int | float) -> str:
    n = int(round(value))
    sign = "-" if n < 0 else ""
    s = str(abs(n))
    if len(s) <= 3:
        return sign + s
    tail = s[-3:]
    head = s[:-3]
    parts = []
    while len(head) > 2:
        parts.append(head[-2:])
        head = head[:-2]
    if head:
        parts.append(head)
    return sign + ",".join(reversed(parts)) + "," + tail


def ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def parse_date_tokens(text: str) -> dt.date | None:
    match = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", text)
    if not match:
        return None
    day = int(match.group(1))
    month = int(match.group(2))
    year = int(match.group(3))
    if year < 100:
        year += 2000
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def format_as_of_date(date_obj: dt.date) -> str:
    return f"{ordinal(date_obj.day)} {date_obj.strftime('%B %Y')}"


def format_surveyed_header_date(date_obj: dt.date) -> str:
    return f"{date_obj.day}-{date_obj.month}-{str(date_obj.year)[-2:]}"


def format_daily_chart_label(date_obj: dt.date) -> str:
    return f"{ordinal(date_obj.day)} {date_obj.strftime('%b')}"


def survey_bar_class(percent: float) -> str:
    if percent >= 10:
        return "high"
    if percent >= 7:
        return "medium"
    return "low"


def approval_bar_class(percent: float) -> str:
    if percent >= 15:
        return "high"
    if percent >= 10:
        return "medium"
    return "low"


def survey_badge_class(percent: float) -> str:
    if percent >= 10:
        return "success"
    if percent >= 7:
        return "warning"
    return "danger"


def approval_badge_class(percent: float) -> str:
    if percent >= 20:
        return "success"
    if percent >= 10:
        return "warning"
    return "danger"


def detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    for i, row in enumerate(rows):
        cells = [normalize_text(cell) for cell in row if normalize_text(cell)]
        if not cells:
            continue
        joined = " | ".join(cells)
        if (
            ("sub district" in joined or "tehsil" in joined)
            and "uploaded village" in joined
            and "uploaded plots" in joined
            and "surveyed" in joined
            and "approved" in joined
        ):
            return i
    raise ValueError("Could not detect header row in Excel sheet.")


def map_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    col: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = normalize_text(cell)
        if not h:
            continue
        if ("sub district" in h or "tehsil" in h) and "name" in h:
            col["name"] = idx
        elif "targeted plots" in h:
            col["targeted_plots"] = idx
        elif "uploaded village" in h:
            col["uploaded_villages"] = idx
        elif "uploaded plots" in h:
            col["uploaded_plots"] = idx
        elif "required target plots" in h:
            col["daily_target"] = idx
        elif "surveyed plots on" in h:
            col["surveyed_today"] = idx
        elif "total surveyed plots" in h:
            col["total_surveyed"] = idx
        elif "percentage" in h and "surveyed" in h and "approved" not in h:
            col["survey_percent"] = idx
        elif "approved by supervisor" in h:
            col["approved"] = idx
        elif "percentage" in h and "approved" in h:
            col["approval_percent"] = idx
        elif "total private surveyors" in h:
            col["total_surveyors"] = idx
        elif "surveyors in fields on" in h:
            col["in_field"] = idx

    required = [
        "name",
        "uploaded_villages",
        "uploaded_plots",
        "daily_target",
        "surveyed_today",
        "total_surveyed",
        "survey_percent",
        "approved",
        "approval_percent",
        "total_surveyors",
        "in_field",
    ]
    missing = [k for k in required if k not in col]
    if missing:
        raise ValueError(f"Missing expected columns in Excel header: {', '.join(missing)}")
    return col


def extract_report_date(rows: list[tuple[Any, ...]], header_row: tuple[Any, ...], col_map: dict[str, int]) -> dt.date:
    for row in rows[:5]:
        for cell in row:
            if isinstance(cell, dt.datetime):
                return cell.date()
            if isinstance(cell, dt.date):
                return cell
            candidate = parse_date_tokens(str(cell)) if cell is not None else None
            if candidate:
                return candidate

    surveyed_header = str(header_row[col_map["surveyed_today"]])
    candidate = parse_date_tokens(surveyed_header)
    if candidate:
        return candidate

    raise ValueError("Could not determine report date from Excel.")


def read_excel_data(excel_path: Path) -> dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = detect_header_row(rows)
    header_row = rows[header_idx]
    col = map_columns(header_row)
    report_date = extract_report_date(rows, header_row, col)

    subdivisions: list[dict[str, Any]] = []
    totals: dict[str, Any] | None = None

    for row in rows[header_idx + 1 :]:
        name_cell = row[col["name"]] if col["name"] < len(row) else None
        if name_cell is None or str(name_cell).strip() == "":
            continue
        raw_name = str(name_cell).strip()
        key = normalize_name(raw_name)

        row_data = {
            "name": display_name(raw_name),
            "uploaded_villages": to_int(row[col["uploaded_villages"]]),
            "uploaded_plots": to_int(row[col["uploaded_plots"]]),
            "daily_target": to_int(row[col["daily_target"]]),
            "surveyed_today": to_int(row[col["surveyed_today"]]),
            "total_surveyed": to_int(row[col["total_surveyed"]]),
            "survey_percent": to_percent(row[col["survey_percent"]]),
            "approved": to_int(row[col["approved"]]),
            "approval_percent": to_percent(row[col["approval_percent"]]),
            "total_surveyors": to_int(row[col["total_surveyors"]]),
            "in_field": to_int(row[col["in_field"]]),
        }

        if key == "total":
            totals = row_data
        else:
            subdivisions.append(row_data)

    if not subdivisions:
        raise ValueError("No subdivision rows found in Excel data.")

    by_name = {row["name"]: row for row in subdivisions}
    ordered = [by_name[name] for name in CANONICAL_ORDER if name in by_name]
    extras = [row for row in subdivisions if row["name"] not in CANONICAL_ORDER]
    extras.sort(key=lambda x: x["name"])
    subdivisions = ordered + extras

    if totals is None:
        totals = {
            "name": "Total",
            "uploaded_villages": sum(r["uploaded_villages"] for r in subdivisions),
            "uploaded_plots": sum(r["uploaded_plots"] for r in subdivisions),
            "daily_target": sum(r["daily_target"] for r in subdivisions),
            "surveyed_today": sum(r["surveyed_today"] for r in subdivisions),
            "total_surveyed": sum(r["total_surveyed"] for r in subdivisions),
            "approved": sum(r["approved"] for r in subdivisions),
            "total_surveyors": sum(r["total_surveyors"] for r in subdivisions),
            "in_field": sum(r["in_field"] for r in subdivisions),
        }
        totals["survey_percent"] = round((totals["total_surveyed"] / totals["uploaded_plots"]) * 100, 2) if totals["uploaded_plots"] else 0.0
        totals["approval_percent"] = round((totals["approved"] / totals["total_surveyed"]) * 100, 2) if totals["total_surveyed"] else 0.0

    return {
        "report_date": report_date,
        "subdivisions": subdivisions,
        "totals": totals,
    }


def build_stats_html(totals: dict[str, Any]) -> str:
    return f"""    <div class="stats-grid">
        <div class="stat-card">
            <div class="stat-label">Total Uploaded Plots</div>
            <div class="stat-value">{format_indian_number(totals["uploaded_plots"])}</div>
            <div class="stat-subtext">Across all sub-divisions</div>
        </div>
        <div class="stat-card success">
            <div class="stat-label">Total Surveyed</div>
            <div class="stat-value">{format_indian_number(totals["total_surveyed"])}</div>
            <div class="stat-subtext">{totals["survey_percent"]:.2f}% completion</div>
        </div>
        <div class="stat-card warning">
            <div class="stat-label">Approved by Supervisor</div>
            <div class="stat-value">{format_indian_number(totals["approved"])}</div>
            <div class="stat-subtext">{totals["approval_percent"]:.2f}% of surveyed</div>
        </div>
        <div class="stat-card danger">
            <div class="stat-label">Daily Target Required</div>
            <div class="stat-value">{format_indian_number(totals["daily_target"])}</div>
            <div class="stat-subtext">To complete by 31st March</div>
        </div>
    </div>"""


def build_subdivision_cards_html(rows: list[dict[str, Any]]) -> str:
    blocks: list[str] = ['    <div class="subdivision-grid">']
    for row in rows:
        survey_class = survey_bar_class(row["survey_percent"])
        approval_class = approval_bar_class(row["approval_percent"])
        blocks.append(
            f"""        <div class="subdivision-card">
            <h3>{row["name"]}</h3>
            <div class="progress-section">
                <div class="progress-label">
                    <span>Survey Progress</span>
                    <span><strong>{row["survey_percent"]:.2f}%</strong></span>
                </div>
                <div class="progress-bar">
                    <div class="progress-fill {survey_class}" style="width: {row["survey_percent"]:.2f}%;"></div>
                </div>
            </div>
            <div class="progress-section">
                <div class="progress-label">
                    <span>Approval Rate</span>
                    <span><strong>{row["approval_percent"]:.2f}%</strong></span>
                </div>
                <div class="progress-bar">
                    <div class="progress-fill {approval_class}" style="width: {row["approval_percent"]:.2f}%;"></div>
                </div>
            </div>
            <div class="detail-grid">
                <div class="detail-item">
                    <span class="detail-label">Total Plots</span>
                    <span class="detail-value">{format_indian_number(row["uploaded_plots"])}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Surveyed Plots</span>
                    <span class="detail-value">{format_indian_number(row["total_surveyed"])}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Approved</span>
                    <span class="detail-value">{format_indian_number(row["approved"])}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">Surveyors</span>
                    <span class="detail-value">{row["total_surveyors"]}</span>
                </div>
                <div class="detail-item">
                    <span class="detail-label">In Field</span>
                    <span class="detail-value">{row["in_field"]}</span>
                </div>
            </div>
        </div>"""
        )
    blocks.append("    </div>")
    return "\n".join(blocks)


def build_table_body_html(rows: list[dict[str, Any]], totals: dict[str, Any]) -> str:
    lines = ["            <tbody>"]
    for row in rows:
        survey_badge = survey_badge_class(row["survey_percent"])
        approval_badge = approval_badge_class(row["approval_percent"])
        lines.append(
            f"""                <tr>
                    <td><strong>{row["name"]}</strong></td>
                    <td>{row["uploaded_villages"]}</td>
                    <td>{format_indian_number(row["uploaded_plots"])}</td>
                    <td>{format_indian_number(row["daily_target"])}</td>
                    <td>{format_indian_number(row["surveyed_today"])}</td>
                    <td>{format_indian_number(row["total_surveyed"])}</td>
                    <td><span class="badge {survey_badge}">{row["survey_percent"]:.2f}%</span></td>
                    <td>{format_indian_number(row["approved"])}</td>
                    <td><span class="badge {approval_badge}">{row["approval_percent"]:.2f}%</span></td>
                    <td>{row["total_surveyors"]}</td>
                    <td>{row["in_field"]}</td>
                </tr>"""
        )

    lines.append(
        f"""                <tr>
                    <td><strong>Total</strong></td>
                    <td><strong>{totals["uploaded_villages"]}</strong></td>
                    <td><strong>{format_indian_number(totals["uploaded_plots"])}</strong></td>
                    <td><strong>{format_indian_number(totals["daily_target"])}</strong></td>
                    <td><strong>{format_indian_number(totals["surveyed_today"])}</strong></td>
                    <td><strong>{format_indian_number(totals["total_surveyed"])}</strong></td>
                    <td><span class="badge {survey_badge_class(totals["survey_percent"])}">{totals["survey_percent"]:.2f}%</span></td>
                    <td><strong>{format_indian_number(totals["approved"])}</strong></td>
                    <td><span class="badge {approval_badge_class(totals["approval_percent"])}">{totals["approval_percent"]:.2f}%</span></td>
                    <td><strong>{totals["total_surveyors"]}</strong></td>
                    <td><strong>{totals["in_field"]}</strong></td>
                </tr>"""
    )
    lines.append("            </tbody>")
    return "\n".join(lines)


def replace_between(text: str, start_marker: str, end_marker: str, replacement: str) -> str:
    start_idx = text.find(start_marker)
    if start_idx == -1:
        raise ValueError(f"Start marker not found: {start_marker}")
    end_idx = text.find(end_marker, start_idx)
    if end_idx == -1:
        raise ValueError(f"End marker not found after start marker: {end_marker}")
    return text[:start_idx] + replacement + text[end_idx:]


def replace_array_const(text: str, const_name: str, values_js: str) -> str:
    pattern = re.compile(rf"const {re.escape(const_name)} = \[[^\]]*];")
    if not pattern.search(text):
        raise ValueError(f"JS array constant not found: {const_name}")
    return pattern.sub(f"const {const_name} = [{values_js}];", text, count=1)


def update_html(index_path: Path, data: dict[str, Any]) -> None:
    html = index_path.read_text(encoding="utf-8")

    report_date = data["report_date"]
    as_of = format_as_of_date(report_date)
    surveyed_header = format_surveyed_header_date(report_date)
    daily_label = format_daily_chart_label(report_date)
    rows = data["subdivisions"]
    totals = data["totals"]

    stats_html = build_stats_html(totals)
    cards_html = build_subdivision_cards_html(rows)
    tbody_html = build_table_body_html(rows, totals)

    html = re.sub(
        r"<p>Ludhiana District \| Data as of [^<]+</p>",
        f"<p>Ludhiana District | Data as of {as_of}</p>",
        html,
        count=1,
    )

    html = replace_between(
        html,
        '    <div class="stats-grid">',
        '\n\n    <h2 style="margin: 30px 0 20px 0; padding-left: 10px;">',
        stats_html,
    )

    html = replace_between(
        html,
        '    <div class="subdivision-grid">',
        '\n\n    <!-- Charts: survey, approval, surveyors, daily progress -->',
        cards_html,
    )

    html = re.sub(
        r"<th>Surveyed \([^)]+\)</th>",
        f"<th>Surveyed ({surveyed_header})</th>",
        html,
        count=1,
    )

    tbody_start = html.find("            <tbody>")
    if tbody_start == -1:
        raise ValueError("Could not find <tbody> in table.")
    tbody_end = html.find("            </tbody>", tbody_start)
    if tbody_end == -1:
        raise ValueError("Could not find </tbody> in table.")
    tbody_end += len("            </tbody>")
    html = html[:tbody_start] + tbody_html + html[tbody_end:]

    subdivisions = [row["name"] for row in rows]
    survey_pcts = [f'{row["survey_percent"]:.2f}' for row in rows]
    approval_pcts = [f'{row["approval_percent"]:.2f}' for row in rows]
    total_surveyors = [str(row["total_surveyors"]) for row in rows]
    in_field = [str(row["in_field"]) for row in rows]
    daily_progress = [str(row["surveyed_today"]) for row in rows]

    html = replace_array_const(html, "subdivisions", ", ".join(f"'{name}'" for name in subdivisions))
    html = replace_array_const(html, "surveyPercentages", ", ".join(survey_pcts))
    html = replace_array_const(html, "approvalPercentages", ", ".join(approval_pcts))
    html = replace_array_const(html, "totalSurveyors", ", ".join(total_surveyors))
    html = replace_array_const(html, "inField", ", ".join(in_field))
    html = replace_array_const(html, "dailyProgress", ", ".join(daily_progress))

    html = re.sub(
        r"label: 'Plots Surveyed \([^']+\)'",
        f"label: 'Plots Surveyed ({daily_label})'",
        html,
        count=1,
    )

    index_path.write_text(html, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Update dashboard HTML from Excel data.")
    parser.add_argument("excel_path", type=Path, help="Path to the Excel (.xlsx) file")
    parser.add_argument(
        "--index",
        type=Path,
        default=Path("index.html"),
        help="Path to dashboard HTML file (default: index.html)",
    )
    args = parser.parse_args()

    if not args.excel_path.exists():
        raise SystemExit(f"Excel file not found: {args.excel_path}")
    if not args.index.exists():
        raise SystemExit(f"HTML file not found: {args.index}")

    data = read_excel_data(args.excel_path)
    update_html(args.index, data)
    print(f"Updated {args.index} from {args.excel_path}")


if __name__ == "__main__":
    main()
